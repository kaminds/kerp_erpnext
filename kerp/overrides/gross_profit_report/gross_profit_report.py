import frappe
from erpnext.accounts.report.gross_profit.gross_profit import execute as gp_execute
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


# Columns this export actually needs, in display order. Built for
# filters.group_by == "Invoice" (the mode with parent/item indent rows) —
# "sales_invoice" only exists as a fieldname in that mode, and Cost Center
# isn't available at all when grouped by Invoice (ERPNext's own report only
# carries cost_center through to row data when grouped by Cost Center).
REQUIRED_FIELDNAMES = [
    "posting_date",
    "sales_invoice",
    "customer",
    "qty",
    "avg._selling_rate",
    "valuation_rate",
    "selling_amount",
    "buying_amount",
    "gross_profit",
    "gross_profit_%",
]


@frappe.whitelist()
def export_gross_profit_xlsx(filters=None):
    # matches the underlying "Gross Profit" report's own role restriction —
    # @frappe.whitelist() alone doesn't gate this by role, so without this a
    # logged-in user with no access to the report itself could still hit
    # this endpoint directly and pull margin data straight out of it
    frappe.only_for(["Accounts Manager", "Accounts User"], message=True)

    filters = (
        frappe.parse_json(filters) if isinstance(filters, str) else (filters or {})
    )
    filters = frappe._dict(filters)

    # everything below (REQUIRED_FIELDNAMES, the parent/item indent styling,
    # the item-name substitution) is built specifically for Invoice grouping
    # — other group-by modes don't have "sales_invoice" as a fieldname at
    # all and have no indent structure, so they'd silently break rather than
    # just produce a wrong-looking file
    group_by = filters.get("group_by") or "Invoice"
    if group_by != "Invoice":
        frappe.throw(
            frappe._("This export only supports Group By \"Invoice\" (got \"{0}\").").format(group_by)
        )

    columns, data = gp_execute(filters)
    columns_by_fieldname = {c.get("fieldname"): c for c in columns}
    visible_cols = [
        columns_by_fieldname[f] for f in REQUIRED_FIELDNAMES if f in columns_by_fieldname
    ]
    # column carrying the item-row indent — was assumed to be column 1 back
    # when "sales_invoice" was the first column; now located by fieldname so
    # it still tracks correctly regardless of REQUIRED_FIELDNAMES' order
    invoice_col_index = next(
        (j for j, c in enumerate(visible_cols, start=1) if c.get("fieldname") == "sales_invoice"),
        None,
    )

    company = filters.get("company") or ""
    currency = (
        frappe.get_cached_value("Company", company, "default_currency")
        if company
        else "USD"
    )
    symbol = frappe.db.get_value("Currency", currency, "symbol") or currency

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gross Profit"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    band_fill = PatternFill("solid", fgColor="F2F6FA")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    total_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    sub_font = Font(italic=True, size=10, color="666666")

    n_cols = len(visible_cols)

    # Title block
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1, f"Gross Profit Report — {company}").font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(
        2,
        1,
        f"Period: {filters.get('from_date')} to {filters.get('to_date')}   |   Generated: {frappe.utils.now_datetime().strftime('%Y-%m-%d %H:%M')}",
    ).font = sub_font

    header_row = 4
    for j, col in enumerate(visible_cols, start=1):
        cell = ws.cell(header_row, j, col["label"])
        cell.font, cell.fill, cell.border = header_font, header_fill, border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    col_widths = [len(c["label"]) + 2 for c in visible_cols]

    row_i = header_row + 1
    for row in data:
        is_total = isinstance(row, dict) and "Total" in str(list(row.values())[0])
        # indent 0 = the invoice-level row, indent 1 = its item rows (indent
        # is absent/None on the Total row, so it never matches here)
        is_parent_row = isinstance(row, dict) and row.get("indent") == 0
        for j, col in enumerate(visible_cols, start=1):
            fieldname = col.get("fieldname")
            value = row.get(fieldname) if isinstance(row, dict) else row[j - 1]

            if fieldname == "sales_invoice" and isinstance(row, dict) and row.get("indent") == 1:
                # on item rows this field holds just the item code (that's
                # what ERPNext's own report puts there) — show the item name
                # instead; the parent row keeps the actual invoice number
                value = row.get("item_name") or value

            if col.get("fieldtype") == "Percent" and isinstance(value, (int, float)):
                # gross_profit_% comes back on a 0-100 scale (e.g. 33.3), but
                # Excel's "0.00%" format multiplies the stored value by 100
                # again to render it — without this it showed as 3330.00%
                value = value / 100.0

            cell = ws.cell(row_i, j, value)
            cell.border = border

            if fieldname in ("posting_date", "qty"):
                cell.alignment = Alignment(horizontal="center")

            if col.get("fieldtype") == "Currency":
                cell.number_format = f'"{symbol}"#,##0.00'
            elif col.get("fieldtype") == "Percent":
                cell.number_format = "0.00%"

            if (
                fieldname == "gross_profit"
                and isinstance(value, (int, float))
                and value < 0
            ):
                cell.font = Font(color="C00000", bold=True)
            elif is_parent_row:
                cell.font = Font(bold=True)

            if is_total:
                cell.font, cell.fill = total_font, total_fill
                cell.border = Border(
                    top=Side(style="double"), bottom=thin, left=thin, right=thin
                )
            elif row_i % 2 == 0:
                cell.fill = band_fill

            if isinstance(row, dict) and row.get("indent") == 1 and j == invoice_col_index:
                cell.alignment = Alignment(indent=2)
                ws.row_dimensions[row_i].outline_level = 1

            col_widths[j - 1] = max(col_widths[j - 1], len(str(value or "")) + 2)
        row_i += 1

    for j, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = min(w, 40)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{row_i - 1}"
    ws.freeze_panes = f"A{header_row + 1}"

    # print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_properties.outlinePr.summaryBelow = False

    file_path = f"/tmp/{frappe.generate_hash()[:10]}.xlsx"
    wb.save(file_path)
    with open(file_path, "rb") as f:
        frappe.local.response.filename = "Gross Profit.xlsx"
        frappe.local.response.filecontent = f.read()
        frappe.local.response.type = "download"
